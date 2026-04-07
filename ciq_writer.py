from openpyxl import load_workbook
from openpyxl.styles import Font, PatternFill
import shutil
from collections import defaultdict


def write_ciq(input_file, output_file,
              sb_subnet, vmotion_subnet,
              sb_records, vmotion_records,
              gateway_ip, broadcast_ip,
              vm_config_records=None,
              drs_rules=None):   # ✅ NEW PARAM

    shutil.copy(input_file, output_file)
    wb = load_workbook(output_file)

    # =========================
    # CIQ TAB
    # =========================
    ws = wb.create_sheet("CIQ")

    headers = ["IP Address", "Hostname", "Description", "VLAN Name"]

    ws["A1"] = f"VM_Network_SB VLAN: {sb_subnet}"

    for i, h in enumerate(headers, 1):
        cell = ws.cell(row=2, column=i)
        cell.value = h
        cell.fill = PatternFill(start_color="92D050", fill_type="solid")
        cell.font = Font(bold=True)

    ws.append([gateway_ip, "Gateway", "Default Gateway", "VM_Network_SB"])

    for r in sb_records:
        ws.append([
            r["IP Address"],
            r["Hostname"],
            r["Description"],
            r["VLAN Name"]
        ])

    ws.append([broadcast_ip, "Broadcast", "Broadcast Address", "VM_Network_SB"])

    # =========================
    # VM CONFIGURATION SHEET
    # =========================
    if vm_config_records:

        ws2 = wb.create_sheet("VM Configuration Sheet")

        headers = [
            "VM Name", "CPU", "RAM", "Memory Reservation", "SWAP",
            "SCSI Controllers", "SCSI Type",
            "Hard Disk -1", "Hard Disk -2", "Hard Disk -3",
            "Hard Disk -4", "Hard Disk -5", "Hard Disk -6",  # ✅ extended
            "Network Adapter"
        ]

        for i, h in enumerate(headers, 1):
            cell = ws2.cell(row=1, column=i)
            cell.value = h
            cell.fill = PatternFill(start_color="FFFF00", fill_type="solid")
            cell.font = Font(bold=True)

        # =========================
        # DATA ROWS
        # =========================
        for r in vm_config_records:

            # -------- SCSI --------
            scsi_list = r.get("SCSI Controllers", []) or []

            ids = [str(s.get("id", 0)) for s in scsi_list]

            scsi_text = f"{len(ids)} ({','.join(ids)})" if ids else ""

            # -------- DISKS GROUPING --------
            disks = r.get("Disks", []) or []

            grouped = defaultdict(list)

            for d in disks:
                name = d.get("disk_name", "")
                base = name.split("_disk")[0] if "_disk" in name else name
                grouped[base].append(d)

            disk_cols = []

            for base, items in grouped.items():

                lines = []

                for d in items:
                    name = d.get("disk_name", "")
                    size = d.get("size", 0)
                    ctrl = d.get("controller", 0)
                    scsi = d.get("scsi_id", 0)
                    mode = d.get("disk_mode", "")

                    line = f"{name} {size}GB ({ctrl}:{scsi})"
                    if mode:
                        line += f" {mode}"

                    lines.append(line)

                # multi-line inside one cell
                disk_cols.append("\n".join(lines))

            # limit to 6 columns
            disk_cols = disk_cols[:6]

            while len(disk_cols) < 6:
                disk_cols.append("")

            # -------- WRITE ROW --------
            ws2.append([
                r["VM Name"],
                r["CPU"],
                r["RAM"],
                r["Memory Reservation"],
                r["SWAP"],
                scsi_text,
                r.get("SCSI Type", ""),
                *disk_cols,
                r.get("Adapter", "")
            ])

    # =========================
    # DRS RULES TAB (NEW)
    # =========================
    if drs_rules:

        ws3 = wb.create_sheet("DRS Rules")

        headers = ["Rule Name", "Type", "VM1", "VM2"]

        for i, h in enumerate(headers, 1):
            cell = ws3.cell(row=1, column=i)
            cell.value = h
            cell.fill = PatternFill(start_color="ADD8E6", fill_type="solid")
            cell.font = Font(bold=True)

        for r in drs_rules:
            ws3.append([
                r.get("Rule Name", ""),
                r.get("Type", ""),
                r.get("VM1", ""),
                r.get("VM2", "")
            ])

    wb.save(output_file)

    print(f"✅ Output Generated: {output_file}")